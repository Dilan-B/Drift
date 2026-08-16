"""
Regenerates Drift-changelog.xlsx — the running log of shipped work.

    python tools/build-changelog.py

WHEN TO RUN: at each version bump / App Store submission (see CLAUDE.md).
Not every commit — the .xlsx is binary, so each rewrite is a full new blob in
git history with an undiffable change.

TO UPDATE: append entries to ROWS below, newest LAST, then rerun. Pull the
raw material with:

    git log --since="<last release date>" --date=short \
        --pretty=format:"===%h|%ad|%s%n%b" --no-merges

One row per user-meaningful change, not per commit — a commit touching five
things gets five rows. Columns are:

    Date, Commit, Type, Area, Change, Details, Release, Needs

  Type    Feature | Bug Fix | UI | Infra | Security | Chore | Docs
          (a new value also needs adding to the Summary sheet's `types`
          list, or its rows go uncounted — the assert at the bottom
          catches this)
  Release the version it shipped in, or "Unreleased" while on main only
  Needs   what it takes beyond a JS reload: Native build | Edge deploy |
          SQL migration | "—"

For a BUG FIX, put the root cause in Details, not just the symptom. That is
the part worth having a year from now.

Requires openpyxl (pip install openpyxl).
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Repo root, resolved from this file so the script runs from any cwd.
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "Drift-changelog.xlsx")

# Date, Commit, Type, Area, Change, Details, Release, Needs
ROWS = [
    # ── 2026-07-07 ────────────────────────────────────────────
    ("2026-07-07", "3cbe56d", "Feature", "Onboarding", "Onboarding moved to after signup and shortened",
     "Pre-signup questionnaire was driving people away; profiling questions dropped entirely.", "1.1.2", "—"),
    ("2026-07-07", "3cbe56d", "Feature", "Screen Time", "30 minutes of screen time pre-added for new accounts",
     "Welcome bonus so a new user isn't locked out before earning anything.", "1.1.2", "—"),
    ("2026-07-07", "3cbe56d", "Bug Fix", "Auth / Sync", "Tasks and screen time disappeared after sign out then sign in",
     "In-app sign-in never called fetchTasks (only cold boot did). Added hydrateTasksFromServer on every sign-in.", "1.1.2", "—"),
    ("2026-07-07", "3cbe56d", "Infra", "Scaling", "Scale review for 100 / 1,000 / 10,000+ users",
     "Bounded task fetches (365-day window, 1,000 rows), caching and rate-limit guards to cut Supabase egress.", "1.1.2", "—"),
    ("2026-07-07", "3cbe56d", "Feature", "Auth", "Sign in with Apple", "Added to onboarding (later hidden until Apple config was done).", "1.1.2", "—"),
    ("2026-07-07", "e167393", "Feature", "Billing", "All payments removed — every feature free",
     "Paywall, upgrade links and purchase UI removed; RevenueCat code kept dormant behind a single flag for later re-enable.", "1.1.2", "—"),
    ("2026-07-07", "7bc033f", "Chore", "Auth", "Phone (SMS OTP) sign-in removed", "Reversed after finding SMS auth costs money per message.", "1.1.2", "—"),

    # ── 2026-07-10 ────────────────────────────────────────────
    ("2026-07-10", "d000ebb", "Bug Fix", "Auth", "Cold-start sign-outs from AES key generation race",
     "Encryption key was generated concurrently on launch; serialized it so cached sessions survive a cold start.", "1.1.2", "—"),
    ("2026-07-10", "ee8f346", "Feature", "Tasks", "Async task valuation — no more waiting on the AI spinner",
     "Task appears instantly with provisional credits; the real AI value patches in when the server answers.", "1.1.2", "—"),
    ("2026-07-10", "ee8f346", "Feature", "Tasks", "Non-productive credit cap",
     "Leisure/maintenance tasks capped at 1/5 of duration, enforced server-side so a stale client can't bypass it.", "1.1.2", "—"),

    # ── 2026-07-12 ────────────────────────────────────────────
    ("2026-07-12", "c13b2b6", "Bug Fix", "Auth", "Google users signed out on every cold start",
     "Email-verification check ran for OAuth users, whose cached session lacks email_confirmed_at. OAuth users are provider-verified — check skipped.", "1.1.2", "—"),
    ("2026-07-12", "c13b2b6", "Bug Fix", "Auth", "Session dropped on transient network errors",
     "getSession failures returned null and logged the user out; now the cached session is preserved.", "1.1.2", "—"),
    ("2026-07-12", "c13b2b6", "Chore", "Release", "Version bump to 1.1.2 (build 3)", "Also installed expo-apple-authentication to fix an EXConstants build error.", "1.1.2", "—"),

    # ── 2026-07-13 ────────────────────────────────────────────
    ("2026-07-13", "2f4dd14", "Feature", "Family", "Family accounts — parent / child / personal",
     "Account type chosen at signup and permanent. Parent shares a family code, assigns tasks, approves completions; child joins passwordless with name + code and earns screen time only after approval.", "1.1.3", "SQL migration"),
    ("2026-07-13", "2f4dd14", "Infra", "Backend", "Family backend: schema v13 + 5 edge functions",
     "families/family_members tables, immutable account_type trigger, cross-account RLS helpers, atomic grant_screen_time RPC; join-family, assign/approve/reject-child-task, set-child-app-policy.", "1.1.3", "SQL migration"),
    ("2026-07-13", "7c402b1", "Bug Fix", "Onboarding", "Returning devices could only ever create a personal account",
     "Devices that had onboarded opened straight on sign-in, skipping the account-type picker. 'Sign up' now routes through it.", "1.1.3", "—"),
    ("2026-07-13", "bb166ee", "Feature", "Onboarding", "Back button through the onboarding flow",
     "Auth returns to the account-type picker so a wrong pick is easy to undo.", "1.1.3", "—"),
    ("2026-07-13", "bb166ee", "Feature", "Onboarding", "Keyboard-aware auth screen",
     "Google button, terms and switch-mode link hide while typing so they stop covering the fields; Return on password submits.", "1.1.3", "—"),
    ("2026-07-13", "80722de", "Feature", "Family", "Family profile page with light/dark toggle",
     "Shared sheet from a header avatar in both shells; sign out and delete account moved here.", "1.1.3", "—"),

    # ── 2026-07-14 ────────────────────────────────────────────
    ("2026-07-14", "80722de", "Bug Fix", "UI", "Unreadable white text on green buttons in dark mode",
     "earn.deep is a LIGHT green in dark mode, so hardcoded white labels vanished. Share code / Approve / Done / +Task fixed.", "1.1.3", "—"),
    ("2026-07-14", "32c8026", "Feature", "Family", "Bottom dock with Home / History tabs",
     "History lists approved tasks with minutes and date; pull-to-refresh on both shells.", "1.1.3", "—"),
    ("2026-07-14", "32c8026", "Feature", "Family", "Live updates without refresh or restart",
     "Parent subscribes to family_members realtime so a kid appears the moment they join.", "1.1.3", "SQL migration"),
    ("2026-07-14", "32c8026", "Chore", "Family", "All emojis removed from family UI and notifications", "Also fixed onboarding rendering in a system-font fallback.", "1.1.3", "—"),
    ("2026-07-14", "879fd9f", "Bug Fix", "UI", "Dark-mode contrast on three more modals",
     "BlockedApps (DONE), AICheck (SUBMIT/CLAIM), UsernameSetup (CLAIM USERNAME) had the same invisible-white-label bug.", "1.1.3", "—"),
    ("2026-07-14", "d91fa4a", "Feature", "Family", "App access: child requests + PIN-gated app picker",
     "Child can request a specific app; parent allows/denies. PIN unlocks Apple's native picker so a parent can choose exact apps on the kid's phone.", "1.1.3", "SQL migration"),
    ("2026-07-14", "813723a", "Bug Fix", "Screen Time", "Random screen time appearing in the morning",
     "Daily reset zeroed the server balance but left the offline pending-stats record, which later wrote the stale balance back. Now cleared as part of the reset.", "1.1.3", "—"),
    ("2026-07-14", "813723a", "Bug Fix", "AI Check", "Rate limit reported as 'check your connection'",
     "Misleading error; now reports the rate limit as a rate limit.", "1.1.3", "—"),
    ("2026-07-14", "93a86cb", "Feature", "Release", "Automatic forced updates on new App Store release",
     "Queries Apple's iTunes lookup for the live version and blocks the whole app behind the update gate. Fails open; skipped in dev/Expo Go.", "1.1.3", "—"),
    ("2026-07-14", "f349c2f", "UI", "Theme", "Dark mode redesigned as its own place, not inverted light mode",
     "Real elevation steps, brighter ink, readable tertiary text, luminous accents.", "1.1.3", "—"),
    ("2026-07-14", "6cc7851", "UI", "Theme", "Full 'greenhouse at night' dark treatment",
     "Visibly green botanical ground instead of charcoal, mint glass borders, chartreuse CTAs.", "1.1.3", "—"),
    ("2026-07-14", "363511e", "UI", "Theme", "Glow and aurora treatment app-wide, both modes",
     "Every primary CTA glows; aurora pools behind hero surfaces and onboarding.", "1.1.3", "—"),
    ("2026-07-14", "14f3bf4", "Security", "Dependencies", "Patched undici (high) and js-yaml (moderate)",
     "HTTP header injection / response-queue poisoning / WebSocket DoS; quadratic-complexity DoS. Lockfile-only, non-breaking.", "1.1.3", "—"),
    ("2026-07-14", "a775c34", "Bug Fix", "Tasks", "Recurring tasks resurrected after delete and piled up as duplicates",
     "Client-only fields were stripped on every server fetch, emptying the dedup guard. Added a persisted per-day guard, field rehydration and duplicate collapsing.", "1.1.3", "—"),

    # ── 2026-07-16 ────────────────────────────────────────────
    ("2026-07-16", "1de2ad1", "Chore", "Release", "Version bump to 1.1.3", "App Store submission.", "1.1.3", "—"),
    ("2026-07-16", "a81ebf4", "Bug Fix", "Tasks", "Recurring duplicates fixed properly — in the database",
     "Three previous client-side guards had all failed; live rows showed duplicates every launch. Moved the invariant into Postgres: a partial unique index makes a duplicate insert impossible.", "1.1.4", "SQL migration"),
    ("2026-07-16", "893a465", "UI", "Shield", "Quiet-paper blocked-app screen",
     "Renders in the app's own theme, calm spare voice, fixed leaf icon replacing random flame/bolt/trophy. New setAppearance bridge mirrors dark mode into the shield.", "1.1.4", "Native build"),
    ("2026-07-16", "893a465", "UI", "Drift In", "Drift In setup rebuilt as 'the greenhouse door'",
     "One composed session card — task, length slider, earn preview — on warm paper. Session logic untouched.", "1.1.4", "—"),
    ("2026-07-16", "893a465", "UI", "Profile", "Profile migrated to the organic-editorial system", "Aurora ground, Playfair username, icon chips, theme-aware legal group.", "1.1.4", "—"),

    # ── 2026-07-17 ────────────────────────────────────────────
    ("2026-07-17", "279feaa", "Feature", "Shield", "Rotating shield themes",
     "15 voice variations for blocked apps and 4 for mid-focus, shown at random instead of one fixed message.", "1.1.4", "Native build"),
    ("2026-07-17", "fd19d84", "Chore", "Release", "Version bump to 1.1.4", "App Store submission — last released version.", "1.1.4", "—"),
    ("2026-07-17", "8e29898", "Bug Fix", "Performance", "Grove buttons missed taps; friend → challenge took many presses",
     "All four tabs stay mounted, so every timer tick re-rendered the whole Grove tree and starved the JS thread. Memoized the screens and gave them stable props.", "Unreleased", "—"),
    ("2026-07-17", "8e29898", "Feature", "Challenges", "Live rep counting for exercise challenges",
     "Activated the dormant MoveNet pose pipeline (a stub file had been shadowing it on every build). Camera counts reps and auto-verifies; photo check remains as fallback.", "Unreleased", "Native build"),
    ("2026-07-17", "8e29898", "UI", "Challenges", "Challenge sheet remodelled",
     "Organic-editorial restyle with LIVE badges on pose-tracked exercises and a plain-language stakes card.", "Unreleased", "—"),
    ("2026-07-17", "8e29898", "UI", "Add Task", "Add Task page rebuilt with the input at the bottom",
     "Thumb-reachable bottom dock for the task name.", "Unreleased", "—"),
    ("2026-07-17", "8e29898", "UI", "Drift In", "Info circle added, bottom tagline removed, input moved to bottom dock", "", "Unreleased", "—"),
    ("2026-07-17", "8e29898", "UI", "Theme", "Aurora backdrop extended to Stats and Grove", "", "Unreleased", "—"),
    ("2026-07-17", "8e29898", "Infra", "CI/CD", "GitHub Actions workflow for EAS builds",
     "Manual trigger with platform + profile inputs. Blocked on Apple Developer account credentials.", "Unreleased", "—"),

    # ── 2026-07-18 ────────────────────────────────────────────
    ("2026-07-18", "da10977", "Bug Fix", "Release", "Forced updates never actually fired",
     "Version lookup used app.json's bundle id (com.drift.app) but the signed app is com.sanghani.drift, so iTunes returned nothing. Corrected both.", "Unreleased", "—"),

    # ── 2026-07-19 ────────────────────────────────────────────
    ("2026-07-19", "0f42415", "Bug Fix", "Notifications", "Out-of-time notification spam",
     "Two code paths fired it, and both re-fired on every app reopen. Now latched with persisted flags: one notification per depletion, re-armed when the balance goes positive.", "Unreleased", "—"),
    ("2026-07-19", "0f42415", "Feature", "Tasks", "Location-based task suggestions",
     "Save your own places (gym, office); arriving fires a notification that opens a prefilled, editable confirm sheet. Coordinates never leave the device — no Places API.", "Unreleased", "Native build"),
    ("2026-07-19", "0f42415", "Feature", "Tasks", "Optional calendar sync",
     "Read-only import of today's events as tasks, deduped by event id so re-syncing never duplicates.", "Unreleased", "Native build"),
    ("2026-07-19", "f41ba2c", "Feature", "Tasks", "AI now classifies the task category",
     "Category picker removed entirely. Classified server-side in the same round trip that values the task, against a fixed whitelist validated on both ends.", "Unreleased", "Edge deploy"),
    ("2026-07-19", "f41ba2c", "Feature", "Calendar", "Google Calendar preferred by default",
     "Google-backed calendars detected, badged, sorted first and preselected; device calendars remain switchable.", "Unreleased", "Native build"),
    ("2026-07-19", "f41ba2c", "UI", "Add Task", "AI explanation collapsed into a tappable mark", "Was a permanent block of text; now one icon that reveals it on tap.", "Unreleased", "—"),
    ("2026-07-19", "f41ba2c", "Feature", "Discoverability", "Automatic tasks surfaced in three quiet places",
     "Optional opt-in step at the end of onboarding, a one-line dismissible nudge on Today, and a single line in Add Task. All self-removing once acted on.", "Unreleased", "—"),
    ("2026-07-19", "d03609a", "UI", "Add Task", "Repeat picker collapsed into a disclosure row", "Five always-visible pills replaced by one row that expands into a list.", "Unreleased", "—"),
    ("2026-07-19", "42112c1", "UI", "Add Task", "Estimated earnings preview + quick length chips",
     "Fills the space the category picker left; 15m/30m/1h/2h chips snap to the slider's step.", "Unreleased", "—"),

    # ── 2026-07-20 ────────────────────────────────────────────
    ("2026-07-20", "1677732", "Feature", "Navigation", "The Lab — new fourth tab for behaviour settings",
     "Holds blocked apps/hours, automatic tasks, recurring tasks and tour replay, moved out of Profile. Stats merged into The Grove to make room.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Feature", "Calendar", "Google as the default calendar source",
     "Explicit connect prompt with an iPhone-calendar fallback; auto-import became a setting (default on).", "Unreleased", "Native build"),
    ("2026-07-20", "1677732", "Feature", "Tasks", "96-entry searchable place bank",
     "Replaced the five hardcoded presets; alias matching so 'uni' finds Campus and 'grocery' finds Supermarket.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Feature", "Onboarding", "App selection required on first run, no longer Pro",
     "Gate only applies where blocking is actually possible, so builds without FamilyControls can't be trapped in onboarding.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Feature", "Onboarding", "Skip on task-picking, swipe-back across the flow", "Plus DEV jump buttons gated on __DEV__.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Bug Fix", "Onboarding", "Strictness picker removed — every option behaved identically",
     "iOS Screen Time re-applies shields on ~15-minute granularity, so 1/3/7/15 all behaved like 15.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Bug Fix", "Tasks", "Level-up modal fired on every sign-in",
     "Sign-in set screen='app' while XP was still 0, so hydrating real XP read as a level gain.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Bug Fix", "Drift In", "Screen-time payout capped at 60m instead of capping session length", "XP stays uncapped.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Bug Fix", "Tour", "Add-task spotlight landed on the wrong element",
     "Ref resolved to the parent row under view flattening; Today now scrolls to top before the tour measures.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "Bug Fix", "Backend", "verify-task logged nothing on rejection",
     "A silent 429 reached the client as 'couldn't connect', indistinguishable from being offline. Every rejection path now logs a reason.", "Unreleased", "Edge deploy"),
    ("2026-07-20", "1677732", "Bug Fix", "AI Check", "Client timeout added; stopped blaming the network for server errors", "Also skips AI verification for tasks with no capturable proof, failing open.", "Unreleased", "Edge deploy"),
    ("2026-07-20", "1677732", "UI", "Theme", "Dark canvas to near-black; light CTA green lifted for contrast",
     "#1F3A2A raised to #3A6B4F — 5.7:1 on cream.", "Unreleased", "—"),
    ("2026-07-20", "1677732", "UI", "Add Task", "Slider haptics and visual pulse on step", "Via optional expo-haptics.", "Unreleased", "—"),
    ("2026-07-20", "e0410be", "Docs", "Process", "Hand-off testing notes for the unverified batch",
     "Listed what to verify in risk order; nothing in that batch was device-verified.", "Unreleased", "—"),
    ("2026-07-20", "98be1a6", "Bug Fix", "UI", "Dark-mode card edges invisible on the near-black canvas",
     "Raised ink.border 0.15→0.28 and ink.hairline 0.09→0.18.", "Unreleased", "—"),
    # ── 2026-07-29 · App Store review fixes ───────────────────
    ("2026-07-29", "7360ba8", "Bug Fix", "Release", "\"Update Required\" locked every install out, including the newest TestFlight build",
     "app.json declared version 1.0.0 while the native build was 1.1.4 (MARKETING_VERSION / CFBundleShortVersionString). The force-update gate reads Constants.expoConfig.version, which comes from app.json rather than the native build settings, so every install compared 1.0.0 against the live App Store version and blocked behind ForceUpdateModal — which has no dismiss. Set app.json and package.json to 1.1.4. The two values are independent and nothing keeps them in sync.",
     "Unreleased", "Native build"),
    ("2026-07-29", "7360ba8", "Bug Fix", "Auth", "Sign in with Apple was never shown, causing two Guideline 4.8 rejections",
     "appleSignIn.js was fully implemented but its import in OnboardingScreen.jsx was commented out and the button never rendered, so only Google was offered. We replied to Apple asserting we had it, which is why it bounced twice.",
     "Unreleased", "Native build"),
    ("2026-07-29", "7360ba8", "Bug Fix", "Onboarding", "Rating prompt fired during onboarding — Guideline 5.6.3 rejection",
     "ReviewPromptScreen was triggered by the tutorial's onDone, before the user had done anything. Moved into completeTask(): fires once ever, after the 3rd completed task, gated on a persisted AsyncStorage flag.",
     "Unreleased", "Native build"),
    ("2026-07-29", "7360ba8", "Feature", "Release", "Hidden dev override on the mandatory-update gate",
     "Recovery path for the lockout above — the gate returns before every other screen, so there is nothing else to tap. Seven taps on the sprout (or visible under __DEV__); deliberately not a plain button, which would defeat forcing a real security update. The grant is stored as the version it was issued for, so it self-clears on the next update instead of disabling the gate forever.",
     "Unreleased", "Native build"),
    ("2026-07-29", "387f630", "Docs", "Release", "Terms of Use link added to the App Store description copy",
     "Guideline 3.1.2(c) auto-rejected the submission twice for its absence. The app still ships dormant auto-renewable-subscription code, so Apple's automated check applies subscription rules regardless of the paywall being disabled, and in-app links do not satisfy the scanner — it reads the store listing.",
     "Unreleased", "—"),
    ("2026-07-29", "166eea9", "Infra", "Backend", "revenuecat-webhook recovered into version control",
     "The function was live and ACTIVE on Supabase with no source in the repo — the only copy was on Supabase's servers, so nobody could review or redeploy it. Downloaded via `supabase functions download`.",
     "Unreleased", "—"),
    # ── 2026-08-16 ────────────────────────────────────────────
    ("2026-08-16", "788c3c0", "Bug Fix", "Tasks", "Calendar-imported tasks skipped AI verification entirely",
     "Tasks created from the calendar suggestion sheet were hardcoded aiCheck: false, so tapping one claimed the credit instantly while every manually-created task went through the AI proof check. Now inherits proAccess like any other task.",
     "1.1.6", "Native build"),
    ("2026-08-16", "788c3c0", "Bug Fix", "Auth", "Users signed out at random, not just after updates",
     "safeGetSession called supabase.auth.signOut() whenever an error message matched /invalid.*refresh|refresh.*expired/i. A transient 5xx, or a cold start before the network was ready, matched that regex and destroyed the session before autoRefreshToken could retry. Supabase already emits SIGNED_OUT on a genuinely dead token, so the manual sign-out was both redundant and destructive. NOTE: a second suspected cause (loss of the Keychain AES key, which makes the encrypted session decrypt to garbage) is untouched and still live.",
     "1.1.6", "—"),
    ("2026-08-16", "20020dc", "Feature", "Growth", "Share-a-win sheet after completing a task",
     "Every third task completed in a day offers a share sheet with the task, duration and time earned, plus an App Store link. Built on React Native's built-in Share, so no new dependency.",
     "1.1.6", "Native build"),
    ("2026-08-16", "be399eb", "Feature", "Siri", "Four Siri voice commands via App Intents",
     "Check balance, create task, start a Drift In session, and today's progress. Registered through AppShortcutsProvider so they also surface in Shortcuts. App Intents needs no entitlement and no App Store Connect configuration, unlike the deprecated SiriKit.",
     "1.1.6", "Native build"),
    ("2026-08-16", "4fba037", "Feature", "Onboarding", "Onboarding slide introducing the Siri commands",
     "Fifth how-it-works slide listing the three main utterances, so the feature is discoverable rather than hidden.",
     "1.1.6", "Native build"),
    ("2026-08-16", "4fba037", "UI", "Interaction", "Buttons now give haptic feedback on press",
     "Anim.Pop accepted a `haptic` prop and silently ignored it - expo-haptics was installed but never reached the button layer. Now fires on press-in rather than onPress, so the tick lands at contact instead of reading as lag. Opt-in per call site.",
     "1.1.6", "Native build"),
    ("2026-08-16", "4fba037", "Feature", "Screen Time", "Mid-session nudge while burning earned time",
     "A silent notification (Is this really the move right now?) fires from the DeviceActivity checkpoint, capped at 4 per day. That checkpoint is the only moment iOS tells a third-party app the user is mid-session on a blocked app.",
     "1.1.6", "Native build"),
    ("2026-08-16", "9452c72", "Infra", "Screen Time", "Checkpoint cadence tightened to 5 minutes, with a hard event cap",
     "More checkpoints means more chances to nudge. Apple documents a ~15-minute floor for threshold delivery, but whether it applies to event thresholds is unverified; every 15-minute mark is still a multiple of 5, so the worst case degrades to the previous behaviour. Also a safety fix: an 8-hour balance previously armed 31 events, and if startMonitoring throws, the catch stops ALL monitoring and leaves no enforcement at all. Now capped at 17.",
     "1.1.6", "Native build"),
    ("2026-08-16", "dbd0fe1", "Feature", "Screen Time", "Block screen names the app it is blocking",
     "ShieldConfigurationDataSource is the one place iOS hands a third-party app the identity of the app being opened, and the parameter was being discarded. The shield can now say Is Instagram the move right now? The name is used only to draw that screen - the extension is sandboxed specifically to prevent moving content out, so it is never persisted or transmitted.",
     "1.1.6", "Native build"),
    ("2026-08-16", "a988d5e", "Bug Fix", "Release", "Xcode Cloud could not upload builds - versions hardcoded in Info.plist",
     "ios/Drift/Info.plist pinned CFBundleShortVersionString and CFBundleVersion to literals. Info.plist OVERRIDES the build settings, so Xcode Cloud's auto-incremented build number never reached the binary. The three extensions never hit this because their Info.plists do not declare these keys. Now reads $(MARKETING_VERSION) / $(CURRENT_PROJECT_VERSION), which also removes Info.plist as a fourth place every version bump had to be applied.",
     "1.1.6", "Native build"),
    ("2026-08-16", "a4c6895", "Bug Fix", "Release", "Archive failed - every App Shortcut utterance needs the applicationName token",
     "Apple rejects the build if any utterance omits it. One of thirteen phrases hardcoded Start Drift In. Since the display name is Drift the token renders inline, so the advertised phrasing was preserved.",
     "1.1.6", "Native build"),
    ("2026-08-16", "7b188c2", "Infra", "Release", "ci_post_clone no longer fails when Node is already installed",
     "brew install node exits non-zero when the formula is already present, and set -e turned that into a dead build before anything compiled. Newer Xcode Cloud images ship Node, so this fired intermittently depending on which image the build landed on.",
     "1.1.6", "—"),
    ("2026-08-16", "2c2ced7", "Bug Fix", "Website", "Download links on the marketing site were dead",
     "Two different App Store IDs were in circulation across docs/ (6738963592 and 6746262733) and both 404'd. The correct ID is 6778215875, verified against the real bundle id com.sanghani.drift via the iTunes lookup API. The same audit is still outstanding on driftproductivity.com, which is a separate codebase.",
     "1.1.6", "—"),
    ("2026-08-16", "2c2ced7", "Docs", "Website", "Blog section, AI and Siri articles, sitemap and structured data",
     "Adds a blog index plus posts on AI verification and Siri control, a sitemap.xml and robots.txt (neither existed before), and JSON-LD MobileApplication/Article/FAQPage markup. Published to the GitHub Pages site, NOT to driftproductivity.com - see SITEBLOG.md for that handoff.",
     "1.1.6", "—"),
]

HEADERS = ["Date", "Commit", "Type", "Area", "Change", "Details", "Release", "Needs"]

# ── Styling ──────────────────────────────────────────────────
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3A2A")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TYPE_FILL = {
    "Bug Fix":  PatternFill("solid", fgColor="FCE4E4"),
    "Feature":  PatternFill("solid", fgColor="E4ECE0"),
    "UI":       PatternFill("solid", fgColor="E8EFF1"),
    "Infra":    PatternFill("solid", fgColor="F1F2EE"),
    "Security": PatternFill("solid", fgColor="FFF2CC"),
    "Chore":    PatternFill("solid", fgColor="F5F5F5"),
    "Docs":     PatternFill("solid", fgColor="F5F5F5"),
}

wb = Workbook()

# ── Sheet 1: Summary ─────────────────────────────────────────
s = wb.active
s.title = "Summary"

s["A1"] = "Drift — Changelog"
s["A1"].font = Font(name=FONT, bold=True, size=16, color="1F3A2A")
s["A2"] = f"Every feature, fix and change shipped. {ROWS[0][0]} – {ROWS[-1][0]}."
s["A2"].font = Font(name=FONT, size=10, italic=True, color="6B7A6E")

s["A4"] = "Total changes"
s["A4"].font = Font(name=FONT, bold=True, size=11)
s["B4"] = f"=COUNTA(Changes!E2:E{len(ROWS) + 1})"
s["B4"].font = Font(name=FONT, bold=True, size=11)

s["A6"] = "By type"
s["A6"].font = Font(name=FONT, bold=True, size=12, color="1F3A2A")
types = ["Feature", "Bug Fix", "UI", "Infra", "Security", "Chore", "Docs"]
s["A7"], s["B7"] = "Type", "Count"
for c in ("A7", "B7"):
    s[c].font = HDR_FONT
    s[c].fill = HDR_FILL
for i, t in enumerate(types, start=8):
    s[f"A{i}"] = t
    s[f"A{i}"].font = BODY
    if t in TYPE_FILL:
        s[f"A{i}"].fill = TYPE_FILL[t]
    s[f"B{i}"] = f'=COUNTIF(Changes!$C$2:$C${len(ROWS) + 1},A{i})'
    s[f"B{i}"].font = BODY

s["D6"] = "By area"
s["D6"].font = Font(name=FONT, bold=True, size=12, color="1F3A2A")
areas = sorted({r[3] for r in ROWS})
s["D7"], s["E7"] = "Area", "Count"
for c in ("D7", "E7"):
    s[c].font = HDR_FONT
    s[c].fill = HDR_FILL
for i, a in enumerate(areas, start=8):
    s[f"D{i}"] = a
    s[f"D{i}"].font = BODY
    s[f"E{i}"] = f'=COUNTIF(Changes!$D$2:$D${len(ROWS) + 1},D{i})'
    s[f"E{i}"].font = BODY

s["G6"] = "By release"
s["G6"].font = Font(name=FONT, bold=True, size=12, color="1F3A2A")
rels = ["1.1.2", "1.1.3", "1.1.4", "1.1.6", "Unreleased"]
s["G7"], s["H7"] = "Release", "Count"
for c in ("G7", "H7"):
    s[c].font = HDR_FONT
    s[c].fill = HDR_FILL
for i, r in enumerate(rels, start=8):
    s[f"G{i}"] = r
    s[f"G{i}"].font = BODY
    s[f"H{i}"] = f'=COUNTIF(Changes!$G$2:$G${len(ROWS) + 1},G{i})'
    s[f"H{i}"].font = BODY

note_row = 8 + len(areas) + 2
s[f"A{note_row}"] = "Notes"
s[f"A{note_row}"].font = Font(name=FONT, bold=True, size=12, color="1F3A2A")
notes = [
    "1.1.4 is the last version submitted to the App Store. Everything marked Unreleased is on main but not shipped.",
    "\"Needs\" says what a change requires beyond a normal JS reload: a native rebuild, an edge-function deploy, or SQL run in Supabase.",
    "Native-build items are currently blocked on Apple Developer account access, so they are not device-tested.",
    "Counts on this sheet are formulas over the Changes tab — they update if you edit, add or delete rows there.",
]
for i, n in enumerate(notes, start=note_row + 1):
    s[f"A{i}"] = "• " + n
    s[f"A{i}"].font = Font(name=FONT, size=10, color="4A4A4A")

for col, w in zip("ABCDEFGH", (28, 10, 6, 22, 10, 6, 14, 10)):
    s.column_dimensions[col].width = w

# ── Sheet 2: Changes ─────────────────────────────────────────
d = wb.create_sheet("Changes")
d.append(HEADERS)
for c in range(1, len(HEADERS) + 1):
    cell = d.cell(row=1, column=c)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = BORDER
d.row_dimensions[1].height = 22

for r in ROWS:
    d.append(list(r))

for row in range(2, len(ROWS) + 2):
    typ = d.cell(row=row, column=3).value
    for col in range(1, len(HEADERS) + 1):
        cell = d.cell(row=row, column=col)
        cell.font = BODY
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 6)))
    d.cell(row=row, column=3).fill = TYPE_FILL.get(typ, PatternFill())
    d.cell(row=row, column=3).alignment = Alignment(vertical="top", horizontal="center")
    d.cell(row=row, column=2).alignment = Alignment(vertical="top", horizontal="center")
    d.cell(row=row, column=7).alignment = Alignment(vertical="top", horizontal="center")
    needs = d.cell(row=row, column=8)
    needs.alignment = Alignment(vertical="top", horizontal="center")
    if needs.value and needs.value != "—":
        needs.font = Font(name=FONT, size=10, color="B0764E", bold=True)
    rel = d.cell(row=row, column=7)
    if rel.value == "Unreleased":
        rel.font = Font(name=FONT, size=10, color="B0764E")
    d.row_dimensions[row].height = 40

for col, w in zip("ABCDEFGH", (11, 9, 9, 14, 46, 66, 11, 14)):
    d.column_dimensions[col].width = w

d.freeze_panes = "A2"
d.auto_filter.ref = f"A1:H{len(ROWS) + 1}"

# openpyxl writes formulas with no cached value, and LibreOffice isn't
# available here to compute them. This flag makes Excel / Sheets / Numbers
# recalculate everything the moment the file opens, so the summary counts
# are populated for the reader.
wb.calculation.fullCalcOnLoad = True

wb.save(OUT)
print(f"Wrote {OUT} with {len(ROWS)} rows")

# ── Verify the formula criteria actually match the data ──────
from collections import Counter
t, a, r = Counter(x[2] for x in ROWS), Counter(x[3] for x in ROWS), Counter(x[6] for x in ROWS)
print("\nExpected counts (what the formulas should produce):")
print("  Types:   ", dict(t))
print("  Releases:", dict(r))
print("  Areas:   ", len(a), "distinct")
assert sum(t.values()) == len(ROWS) and sum(r.values()) == len(ROWS)
for k in t:
    assert k in ("Feature", "Bug Fix", "UI", "Infra", "Security", "Chore", "Docs"), f"unlisted type {k}"
for k in r:
    assert k in ("1.1.2", "1.1.3", "1.1.4", "1.1.6", "Unreleased"), f"unlisted release {k}"
print("\nAll type/release values are covered by a summary row - no change is uncounted.")
